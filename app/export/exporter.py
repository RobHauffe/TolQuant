import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict, Optional
import numpy as np

def export_to_excel(file_path: str, data: Dict, is_itt: bool = False):
    """
    Export all data to an Excel file with multiple sheets.
    data structure expected:
    - 'curve_data': pd.DataFrame (mean/SEM per time point)
    - 'summary_stats': pd.DataFrame (metrics per group)
    - 'sample_details': pd.DataFrame (all samples metrics)
    - 'raw_data': pd.DataFrame (original values, with exclusion markers)
    - 'stats_tests': Dict[str, pd.DataFrame] (ANOVA, post-hoc)
    """
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    # Note: metric_labels are now generated in main.py before export
    # This function just writes the data as provided
    # Sheet 1: Curve Data
    if 'curve_data' in data:
        data['curve_data'].to_excel(writer, sheet_name='Curve Data', index=False)
    
    # Sheet 2: Summary Statistics
    if 'summary_stats' in data:
        data['summary_stats'].to_excel(writer, sheet_name='Summary Statistics', index=False)
        
    # Sheet 3: Sample Details
    if 'sample_details' in data:
        data['sample_details'].to_excel(writer, sheet_name='Sample Details', index=False)
        
    # Sheet 4: Raw Data
    if 'raw_data' in data:
        raw_df = data['raw_data']
        raw_df.to_excel(writer, sheet_name='Raw Data', index=False)
        
        # Apply red fill to excluded values
        # (This requires more specific logic if we want to format based on excluded list)
        # We'll pass a 'raw_data_mask' indicating exclusions
        if 'raw_data_mask' in data:
            ws = writer.sheets['Raw Data']
            mask = data['raw_data_mask']
            red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
            
            # mask is expected to be a boolean DataFrame of same shape as raw_df (minus non-data columns)
            # Find the starting column for time point data in raw_df
            # Assuming raw_df columns are: ['Group', 'Sample ID', 'Time_0', 'Time_15', ...]
            start_col = 3 # 1-indexed (C)
            for r_idx, row in enumerate(mask.values):
                for c_idx, is_excluded in enumerate(row):
                    if is_excluded:
                        # +2 because excel is 1-indexed and we skip Group/Sample ID
                        ws.cell(row=r_idx+2, column=c_idx+start_col).fill = red_fill
    
    # Sheet 5: Statistical Tests
    if 'stats_tests' in data:
        start_row = 1
        ws_stats = writer.book.create_sheet('Statistical Tests')
        for test_name, test_df in data['stats_tests'].items():
            ws_stats.cell(row=start_row, column=1, value=test_name).font = Font(bold=True)
            start_row += 1
            if isinstance(test_df, pd.DataFrame):
                for r in dataframe_to_rows(test_df, index=True, header=True):
                    for c_idx, value in enumerate(r, 1):
                        ws_stats.cell(row=start_row, column=c_idx, value=value)
                    start_row += 1
                start_row += 2 # gap between tests
            elif isinstance(test_df, str):
                ws_stats.cell(row=start_row, column=1, value=test_df)
                start_row += 2

    # Formatting
    for sheetname in writer.sheets:
        ws = writer.sheets[sheetname]
        # Bold headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    writer.close()

def generate_summary_tsv(summary_df: pd.DataFrame) -> str:
    """Generate tab-separated string for clipboard."""
    return summary_df.to_csv(sep='\t', index=False)
